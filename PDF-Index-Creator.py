import sys
import time
from enum import Enum
import PyPDF2
import os
import openpyxl
from PySide6.QtCore import Qt, QUrl, QEvent, QPoint, QModelIndex, QCoreApplication, QTimer,  QObject, QSize, \
    QEventLoop, Signal
from PySide6.QtGui import  QIcon, QIntValidator, QPalette,  QCursor, QDesktopServices
from PySide6.QtWebEngineWidgets import QWebEngineView
from PySide6.QtWebEngineCore import QWebEngineSettings
from PySide6.QtWidgets import (
    QApplication,
    QGraphicsDropShadowEffect,
    QHBoxLayout,
    QComboBox,
    QStyle,
    QPushButton,
    QVBoxLayout,
    QWidget,
    QSplitter,
    QFileDialog,
    QMessageBox,
    QSystemTrayIcon,
    QStyleOptionViewItem,
    QLabel

)

from PySide6.QtWidgets import  QDialog, QTableWidgetItem, QHeaderView
from qframelesswindow.webengine import FramelessWebEngineView
from qfluentwidgets import (SplashScreen, InfoBar,
                            InfoBarPosition, FluentWindow,NavigationAvatarWidget,NavigationToolButton,
                            NavigationItemPosition)
from qfluentwidgets import FluentIcon as FIF
from qfluentwidgets import isDarkTheme,TextEdit
from qfluentwidgets import ComboBox
from qfluentwidgets import (FluentIcon,RoundMenu, CommandBar, Action,CaptionLabel,
                            MessageBox,
                            PrimaryPushButton)
from qframelesswindow import FramelessWindow
from TableViewCustom import TableItemDelegate, TableWidget
import re
from qfluentwidgets import LineEdit, PushButton, setTheme, Theme



def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path).replace('\\', '/')


PDFJS = "file:///" + resource_path("pdfjs/web/viewer.html").lstrip("/")


class CustomTableItemDelegate(TableItemDelegate):
    def __init__(self, parent, widgets):
        super().__init__(parent)
        self.widgets = widgets  # Expected format: {QWidgetType: [column_indices]}

    def createEditor(self, parent, option, index):
        for widget_type, columns in self.widgets.items():
            if index.column() in columns:
                if widget_type == ComboBox:
                    return None  # Don't create an editor for ComboBox
                editor = widget_type(parent)
                # Set a solid background color for the editor based on the theme
                if isDarkTheme():
                    current_stylesheet = editor.styleSheet()
                    override_stylesheet = """
                        TextEdit:focus, PlainTextEdit:focus {
                            border-bottom: 1px solid --ThemeColorPrimary;
                            background-color: #353535;
                        }
                        LineEdit:focus[transparent=true] {
                            background: #353535;
                            border-bottom: 1px solid rgba(255, 255, 255, 0.08);
                        }

                        LineEdit[transparent=false]:focus {
                            background: #353535;
                        }                   
                    """
                    print(current_stylesheet)
                    # Append the new style rule to the current stylesheet
                    new_stylesheet = current_stylesheet + override_stylesheet
                    editor.setStyleSheet(new_stylesheet)
                # else:
                # pass
                # editor.setStyleSheet("background-color: white;")
                return editor
        return super().createEditor(parent, option, index)

    def setModelData(self, editor, model, index):
        try:
            for widget_type, columns in self.widgets.items():
                if index.column() in columns and isinstance(editor, widget_type):
                    value = self._get_editor_value(editor)
                    if value is not None:
                        model.setData(index, value, Qt.DisplayRole)
                        table = self.parent()
                        table.resizeRowToContents(index.row())  # Corrected to resize row
                        return
            # Fall back to the default implementation if the conditions above are not met.
            super().setModelData(editor, model, index)
        except Exception as e:
            print(f"An error occurred in setModelData: {str(e)}")
            # Optionally, you might want to log the traceback to understand where the error is coming from
            import traceback
            print(traceback.format_exc())

    def setEditorData(self, editor, index):
        for widget_type, columns in self.widgets.items():
            if index.column() in columns and isinstance(editor, widget_type):
                value = index.data(Qt.DisplayRole)
                self._set_editor_value(editor, value)
                return
        super().setEditorData(editor, index)

    def _get_editor_value(self, editor):
        """Retrieve the value from the editor widget."""
        try:
            if isinstance(editor, TextEdit):
                return editor.toPlainText()
            elif isinstance(editor, LineEdit):
                return editor.text()
            elif isinstance(editor, ComboBox):
                return editor.currentText()
            else:
                print(f"Unhandled editor type: {type(editor)}")
                return None
        except Exception as e:
            print(f"An error occurred in _get_editor_value: {str(e)}")
            return None

    def _set_editor_value(self, editor, value):
        if isinstance(editor, (LineEdit, TextEdit)):
            editor.setText(str(value))
        elif isinstance(editor, ComboBox):
            # Assume values are already in ComboBox or add additional logic
            combo_index = editor.findText(str(value))
            if combo_index >= 0:
                editor.setCurrentIndex(combo_index)

    def initStyleOption(self, option: QStyleOptionViewItem, index: QModelIndex):
        super().initStyleOption(option, index)
        if index.column() != 5:
            return

        if isDarkTheme():
            option.palette.setColor(QPalette.Text, Qt.black)
            option.palette.setColor(QPalette.HighlightedText, Qt.black)
        else:
            option.palette.setColor(QPalette.Text, Qt.white)
            option.palette.setColor(QPalette.HighlightedText, Qt.white)


# noinspection PyUnresolvedReferences


class PDFIndexCreator(QWidget):
    # noinspection PyUnresolvedReferences
    """This class is the main widget of the application. It contains the table and the \
    buttons to add/remove rows/columns.
    #PARAMETERS: QWidget
    #RETURNS: None
    """

    # Nested TableWidget class
    class TableWidget(TableWidget):
        """Custom table functionality."""

        def __init__(self, parent):
            super().__init__(parent)

        def contextMenuEvent(self, event):
            """Show context menu on right-click."""
            item = self.itemAt(event.pos())
            if item:
                self.selectRow(item.row())
                self.createContextMenu(item, event)

        def createContextMenu(self, item, event):
            """Create context menu for table rows."""
            menu = RoundMenu(parent=self)  # Custom context menu

            # Actions for the context menu
            addRowAboveAction = Action(FIF.CARE_UP_SOLID, 'Add Row Above')
            addRowBelowAction = Action(FIF.CARE_DOWN_SOLID, 'Add Row Below')
            deleteRowAction = Action(FIF.DELETE, 'Remove Row')

            # Add actions to the menu
            menu.addAction(addRowAboveAction)
            menu.addAction(addRowBelowAction)
            menu.addSeparator()  # Separator between add and delete actions
            menu.addAction(deleteRowAction)

            # Connect actions to methods
            addRowAboveAction.triggered.connect(lambda: self.parent().add_row_table_context(item.row()))
            addRowBelowAction.triggered.connect(lambda: self.parent().add_row_table_context(item.row() + 1))
            deleteRowAction.triggered.connect(self.parent().remove_row)

            menu.exec(event.globalPos())

    def __init__(self, combine_instance):
        """Constructor of the class. Initializes the table and buttons."""
        super().__init__()
        self.main_window = combine_instance  # Reference to the combine instance
        self.layout = QVBoxLayout()
        self.history = []
        self.future_states = []
        self.current_version = -2
        self.resize(600, 750)

        # Initialize and configure table
        self.table = self.TableWidget(self)
        self.table.setRowCount(0)
        self.table.setColumnCount(6)
        self.table.verticalHeader().hide()
        self.table.setHorizontalHeaderLabels(["Word", "Page", "Notes/Comments", "Book", "Reference", "Icon"])
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().sectionClicked.connect(self.onHeaderClicked)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionsMovable(True)
        self.table.cellDoubleClicked.connect(self.handle_double_click)
        self.table.setWordWrap(True)
        self.table.setTextElideMode(Qt.TextElideMode.ElideNone)

        # Retrieve the current stylesheet of the table
        current_stylesheet = self.table.styleSheet()
        '''
        # Define the new style rules
        override_stylesheet = """
            QHeaderView::section:vertical {
                /* Add your custom styles here, e.g., min-height */
                min-height: 100px;  /* Adjust as needed */
                /* Other styles as needed */
            }
            /* Any other style overrides */
        """

        # Append the new style rule to the current stylesheet
        new_stylesheet = current_stylesheet + override_stylesheet
        '''

        # Hide columns initially
        self.table.setColumnHidden(3, True)  # Book
        self.table.setColumnHidden(4, True)  # Reference
        self.table.setColumnHidden(5, True)  # Icon

        # Initialize delegate for table
        widgets = {TextEdit: [2, 4], ComboBox: [5], LineEdit: [0, 1, 3]}
        self.delegate = CustomTableItemDelegate(self.table, widgets)
        self.table.setItemDelegate(self.delegate)
        # Apply the updated stylesheet to the table
        # self.table.setStyleSheet(new_stylesheet)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        # Column resizing flag
        self.is_column0_resized = False
        self.adjustHeaders()

        # Setup command bar with buttons
        hBoxLayout = QHBoxLayout()
        self.commandBar = CommandBar(self)
        hBoxLayout.addWidget(self.commandBar, 0)
        self.commandBar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.commandBarButton(FluentIcon.ADD_TO, 'Add Row')
        self.commandBarButton(FluentIcon.REMOVE_FROM, 'Remove Row')
        self.commandBarButton(FluentIcon.HISTORY, 'Undo')
        self.commandBarButton(FluentIcon.ROTATE, 'Redo')

        self.commandBar.addSeparator()

        # Add actions for column visibility
        self.book_action = Action(FluentIcon.EDIT, 'Show Book Column', checkable=True)
        self.book_action.triggered.connect(
            lambda isChecked=self.book_action.isChecked(): self.handle_book_column_visibility(isChecked))
        self.commandBar.addAction(self.book_action)

        self.action = Action(FluentIcon.EDIT, 'Show Reference Column', checkable=True)
        self.action.triggered.connect(
            lambda isChecked=self.action.isChecked(): self.handle_reference_column_visibility(isChecked))
        self.commandBar.addAction(self.action)

        self.icon_action = Action(FluentIcon.EDIT, 'Show Icon Column', checkable=True)
        self.icon_action.triggered.connect(
            lambda isChecked=self.icon_action.isChecked(): self.handle_icon_column_visibility(isChecked))
        self.commandBar.addAction(self.icon_action)

        # Finalize layout
        self.layout.addWidget(self.table)
        self.layout.addLayout(hBoxLayout)
        self.setLayout(self.layout)
        self.save_to_history()

        # Connect resizing event
        self.table.horizontalHeader().sectionResized.connect(self.column_resized)

    def commandBarButton(self, icon, text):
        # Define a dictionary to map the text labels to their corresponding methods.
        action_methods = {
            'Add Row': self.add_row,
            'Remove Row': self.remove_row,
            'Undo': self.undo,
            'Redo': self.redo  # ,
            # 'Show Book Column': self.book_visibility_checkbox
        }

        # Create an action object.
        action = Action(icon, text, self)

        # If the text label exists in the dictionary, connect its corresponding method.
        # Otherwise, just print the text.
        if text in action_methods:
            action.triggered.connect(action_methods[text])
        else:
            action.triggered.connect(lambda: print(text))

        # Add the action to the command bar.
        self.commandBar.addAction(action)

    def column_resized(self, logicalIndex, oldSize, newSize):
        if logicalIndex == 0:
            self.is_column0_resized = True

    def onHeaderClicked(self, logicalIndex):
        if logicalIndex == 0:  # If the first header is clicked
            self.table.setSortingEnabled(False)  # Disable sorting for this header for now TESTING
        else:
            self.table.setSortingEnabled(False)  # Disable sorting for other headers

    def adjustHeaders(self):
        header = self.table.horizontalHeader()
        '''
                header = self.table.horizontalHeader()

                # Set the resize mode of the columns
                header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)  # Word
                header.setSectionResizeMode(1, QHeaderView.Fixed)  # Page
                #header.setSectionResizeMode(1, QHeaderView.Fixed)
                #header.resizeSection(1, 10)  # Replace 'desired_width' with the width you want in pixels
                header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Description
                header.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)  # Book
                header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # Reference
                header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # Icon

                #header.resizeSection(2, 500)
        '''
        default_resize_modes = {
            0: QHeaderView.Interactive,  # Word
            1: QHeaderView.Fixed,  # Page
            2: QHeaderView.Stretch,  # Description
            3: QHeaderView.Interactive,  # Book
            4: QHeaderView.Stretch,  # Reference
            5: QHeaderView.ResizeToContents  # Icon
        }

        empty_resize_modes = {
            0: QHeaderView.Interactive,  # Word
            1: QHeaderView.Fixed,  # Page
            2: QHeaderView.Stretch,  # Description
            3: QHeaderView.Fixed,  # Book
            4: QHeaderView.Stretch,  # Reference
            5: QHeaderView.Fixed  # Icon
        }

        # If column 0 has been manually resized, keep it in Interactive mode
        if not self.is_column0_resized:
            header.setSectionResizeMode(0, QHeaderView.Fixed)
        else:
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
            # header.resizeSection(0, max(header.sectionSize(0), 10))  # Use 100 or a suitable width

        if self.table.rowCount() == 0:
            # For each section, resize it to the header's text width
            for i in range(header.count()):
                header_text = header.model().headerData(i, Qt.Horizontal)
                font_metrics = self.fontMetrics()
                header_length = font_metrics.horizontalAdvance(header_text) + 20  # Add some padding
                header.resizeSection(i, header_length)

            # Stretch the last visible column
            for i in reversed(range(header.count())):
                if not self.table.isColumnHidden(i):
                    header.setSectionResizeMode(i, QHeaderView.Stretch)
                    break  # exit loop once we've found and set the last visible column

        else:
            # Use the default resize modes if column has content, otherwise use the empty resize mode
            for i in range(header.count()):
                if self.columnHasContent(i):
                    header.setSectionResizeMode(i, default_resize_modes.get(i, QHeaderView.Fixed))
                else:
                    header.setSectionResizeMode(i, empty_resize_modes.get(i, QHeaderView.Fixed))

    def columnHasContent(self, columnIndex):
        """Check if the given column has any non-empty content."""
        for row in range(self.table.rowCount()):
            item = self.table.item(row, columnIndex)
            if item and item.text().strip():  # Check if item exists and it has non-empty content
                return True
        return False

    def clear_table(self):
        """
        This function clears the table.
        return:        None
        param:         None
        """
        # Clear the row count of the table
        self.table.setRowCount(0)

    def reapply_combo_box_style(self, theme):
        # Define the styles for both themes
        light_theme_style = """
                ComboBox {
            border: none;
            border-radius: 0px;
            padding: 5px 31px 6px 11px;
            color: black;
            background-color: transparent;
            text-align: left;
        }
        ComboBox:pressed, QComboBox:on {
            border-radius: 5px;
        }
        ComboBox:disabled {
            color: rgba(0, 0, 0, 0.36);
            background: rgba(249, 249, 249, 0.3);
            border: 1px solid rgba(0, 0, 0, 0.06);
            border-bottom: 1px solid rgba(0, 0, 0, 0.06);
        }
                """

        dark_theme_style = """
                ComboBox {
                border: none;
                border-radius: 0px;
                padding: 5px 31px 6px 11px;
                color: white;
                background-color: transparent;
                text-align: left;
            }
            ComboBox:pressed, QComboBox:on {
                border-radius: 5px;
            }
            ComboBox:disabled {
                color: rgba(255, 255, 255, 0.36);
                background: rgba(50, 50, 50, 0.3);
                border: 1px solid rgba(255, 255, 255, 0.06);
                border-bottom: 1px solid rgba(255, 255, 255, 0.06);
            }
                """

        # Choose the style based on the theme
        combo_box_style = light_theme_style if theme == Theme.LIGHT else dark_theme_style

        # Update the style for all combo boxes
        for row in range(self.table.rowCount()):
            icon_combobox = self.table.cellWidget(row, 5)
            if icon_combobox is not None:
                icon_combobox.setStyleSheet(combo_box_style)

    def populate_icon_column(self, row):
        """
        This function populates the icon column with a combobox.
        param row:         The row to populate
        return:        None
        """
        # icon_combobox = QComboBox(self)
        icon_combobox = ComboBox(self)

        icon_combobox.addItems(
            [
                "None",
                "cross-platform",
                "Android",
                "cloud icon with a server",
                "dollar (for paid asset)",
                "iOS",
                "Linux",
                "Mac OS",
                "paper note with a pen",
                "USB key",
                "Python script",
                "general-purpose script",
                "red question mark",
                "Solaris",
                "Web-related asset",
                "Windows",
            ]
        )  # Replace this with your actual icon names
        # icon_combobox.setStyleSheet(combo_box_style)
        # Install the event filter for the ComboBox
        icon_combobox.installEventFilter(self.delegate)

        icon_combobox.currentIndexChanged.connect(
            lambda: self.handle_icon_change(row, icon_combobox)
        )
        self.table.setCellWidget(row, 5, icon_combobox)
        # self.main_window.current_theme)
        self.reapply_combo_box_style(self.main_window.current_theme)
        # self.table.setCellWidget(
        #    row, 0, icon_combobox
        # )  # Use 5 instead of 3 for the Icon column

    def handle_icon_change(self, row, combobox):
        """
        DESCRIPTION: This function handles the icon change.
        param row:     The row that was changed
        param combobox:  The combobox that was changed
        return:    None
        """
        selected_icon = combobox.currentText()  # Get current text from the combobox

        # Check if an item exists at the specified cell
        item = self.table.item(row, 5)

        if item:
            item.setText(selected_icon)
        else:
            # Create a new QTableWidgetItem if none exists
            self.table.setItem(row, 5, QTableWidgetItem(selected_icon))

        self.save_to_history()

    def handle_reference_column_visibility(self, state):
        """
        DESCRIPTION: This function handles the reference column visibility.
        param state:   The state of the checkbox
        return:        None
        """
        QApplication.setOverrideCursor(Qt.WaitCursor)  # Block GUI updates
        self.table.setColumnHidden(4, not state)  # Toggle the "Reference" column visibility based on state

        # Use 'self.action' here because 'action' is not defined in this scope
        if state:
            self.action.setText('Hide Reference Column')  # Update text to reflect the action of hiding the column
        else:
            self.action.setText('Show Reference Column')  # Update text to reflect the action of showing the column

        QApplication.restoreOverrideCursor()  # Allow GUI updates again

    def handle_book_column_visibility(self, state):
        """
        Handle the visibility of the "Book" column in a QTableWidget based on the given state.
        param state: The state of the "Book" column visibility.
        return: None
        """
        QApplication.setOverrideCursor(Qt.WaitCursor)  # Block GUI updates
        self.table.setColumnHidden(3, not state)  # Toggle the "Book" column visibility

        # Update the text of the QAction for the "Book" column
        if state:
            self.book_action.setText('Hide Book Column')
        else:
            self.book_action.setText('Show Book Column')

        QApplication.restoreOverrideCursor()  # Allow GUI updates again

    def handle_icon_column_visibility(self, state):
        """
        DESCRIPTION: This function handles the icon column visibility.
        param state:       The state of the checkbox
        return:        None
        """
        QApplication.setOverrideCursor(Qt.WaitCursor)  # Block GUI updates
        self.table.setColumnHidden(5, not state)  # Toggle the "Icon" column visibility

        # Update the text of the QAction for the "Icon" column
        if state:
            self.icon_action.setText('Hide Icon Column')
        else:
            self.icon_action.setText('Show Icon Column')

        QApplication.restoreOverrideCursor()  # Allow GUI updates again
        QCoreApplication.processEvents()  # Process any pending GUI events
        self.stretch_table()  # Adjust table columns

    def save_state(self):
        """
        DESCRIPTION: This function saves the current state of the table to the history list.
        return:        None
        """
        # Save the current state to the history
        self.save_to_history()

    def save_to_history(self):
        """
        DESCRIPTION: This function saves the current state of the table to the history list.
        return:    None
        """
        # Only save to history if this is not an undo/redo operation
        if not self.main_window.last_op_undo_redo:
            # If the history list is already at capacity, remove the oldest item
            if len(self.history) >= 10:  # We only want to keep 5 versions of the table
                self.history.pop(0)
            # Add the current table data to the history list
            self.history.append(self.get_table_data())
            # Clear future_states every time you save to history
            self.future_states = (
                []
            )  # clear future_states every time you save to history
            # Set the current version to the last item in the history list
            self.current_version = len(self.history) - 1

    def undo(self):
        """
        DESCRIPTION: This function undoes the last operation.
        return:        None
        """
        # Set the last operation to undo
        self.main_window.last_op_undo_redo = True
        # Increment the update counter
        self.main_window.update_counter = 0
        # If the current version is greater than 0
        if self.current_version > 0:
            # Save the current state to the future states
            self.future_states.append(self.history[self.current_version])
            # Undo operation
            self.current_version -= 1
            self.set_table_data(self.history[self.current_version])
        # Reset the update counter
        self.main_window.update_counter = 0
        # Set the last operation to undo
        self.main_window.last_op_undo_redo = False
        # self.stretch_table()
        self.reapply_combo_box_style(self.main_window.current_theme)

    def redo(self):
        """
        DESCRIPTION: This function redoes the last operation.
        return:        None
        """
        # Set the last operation to redo
        self.main_window.last_op_undo_redo = True
        # Increment the update counter
        self.main_window.update_counter = 0
        # If the future states are not empty
        if self.future_states:
            # Redo operation
            future_state = self.future_states.pop()
            self.set_table_data(future_state)
            self.current_version = len(self.history) - 1
        # Reset the update counter
        self.main_window.update_counter = 0
        # Set the last operation to redo
        self.main_window.last_op_undo_redo = False
        # self.stretch_table()
        self.reapply_combo_box_style(self.main_window.current_theme)

    def add_row_table_context(self, position=None):
        # print("Before adding:", self.table.rowCount())  # Debug print
        """
        DESCRIPTION: This function adds a new row to the table.
        :param position: Position where to add the new row. If None, adds to the end.
        return:    None
        """
        # Store the current number of rows
        previous_row_count = self.table.rowCount()

        # If no position is specified, insert a new row at the end.
        if position is None:
            position = self.table.rowCount()
            # self.table.insertRow(row_position)
            # position = previous_row_count

        self.table.insertRow(position)

        # Insert dummy value to new row
        for col in range(self.table.columnCount()):
            self.table.setItem(position, col, QTableWidgetItem(""))

        # Check if any new rows were added
        if previous_row_count < self.table.rowCount():
            # Populate the "Icon" column for the new row
            self.populate_icon_column(position)

        # Save the history
        # print("After adding:", self.table.rowCount())  # Debug print
        self.table.clearSelection()
        self.save_to_history()

    def add_row(self):
        """
        DESCRIPTION: This function adds a new row to the table.
        return:    None
        """
        # Store the current number of rows
        previous_row_count = self.table.rowCount()

        # Insert a new row
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)

        # Insert dummy value to new row
        for col in range(self.table.columnCount()):
            self.table.setItem(row_position, col, QTableWidgetItem(""))

        # Check if any new rows were added
        if previous_row_count < self.table.rowCount():
            # Populate the "Icon" column for the new row
            self.populate_icon_column(row_position)

        # Save the history
        self.table.clearSelection()
        self.save_to_history()
        # self.indexCreator.save_state()

    def remove_row(self):
        """
        DESCRIPTION: This function removes a row from the table.
        return:        None
        """
        # Check if there are selected items
        if len(self.table.selectedItems()) > 0:
            # Remove the selected item
            self.table.removeRow(self.table.selectedItems()[0].row())

            # Clear the selection after removing the row
            self.table.clearSelection()
        self.save_to_history()

    def handle_double_click(self, row, column):
        """
        DESCRIPTION: This function handles the double click event.
        param row:         The row of the cell that was double-clicked
        param column:      The column of the cell that was double-clicked
        return:            None
        """

        # Get the item from column 2 of the double-clicked row
        item_in_column_2 = self.table.item(row, 5)

        # Print the value of the item in column 2
        if item_in_column_2:
            pass
            # print(item_in_column_2.text())

        item = self.table.item(row, column)
        if item:
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.table.editItem(item)
            self.save_to_history()

    def get_table_data(self):
        """
        DESCRIPTION: This function gets the data from the table.
        return:    A list of lists containing the data from the table
        """
        # Get the number of rows in the table
        rows = self.table.rowCount()
        # Get the number of columns in the table
        columns = self.table.columnCount()
        # Get the column names from the table
        column_names = [
            self.table.horizontalHeaderItem(i).text()
            if self.table.horizontalHeaderItem(i)
            else ""
            for i in range(columns)
        ]
        # Create an empty list to store the data
        data = []
        # Iterate through each row in the table
        for row in range(rows):
            # Create an empty list to store the row data
            row_data = []
            # Iterate through each column in the table
            for col in range(columns):
                # If this is the icon column
                if col == 5:  # If this is the icon column
                    # Get the current index of the icon combobox
                    icon_combobox = self.table.cellWidget(row, col)
                    # Append the current index to the row data
                    row_data.append(icon_combobox.currentIndex())
                # If this is not the icon column
                else:
                    # Append the text of the current cell to the row data
                    row_data.append(
                        self.table.item(row, col).text()
                        if self.table.item(row, col)
                        else None
                    )
            # Append the row data to the data list
            data.append(row_data)
        # Return the data
        return rows, columns, column_names, data

    def set_table_data(self, data):
        """
        DESCRIPTION: This function sets the data in the table.
        param data:        The data to set in the table
        return:        None
        """

        # Get the number of rows and columns from the data
        rows, columns, column_names, data = data
        # Set the number of rows and columns in the table
        self.table.setRowCount(rows)
        self.table.setColumnCount(columns)
        # Iterate through the columns and set the header item for each
        for col in range(columns):
            self.table.setHorizontalHeaderItem(col, QTableWidgetItem(column_names[col]))
        # Iterate through the rows and set the data for each
        for row in range(rows):
            for col in range(columns):
                # If the column is the icon column
                if col == 5:  # If this is the icon column
                    # Create a new combo box
                    # icon_combobox = QComboBox(self)
                    icon_combobox = ComboBox(self)
                    # Add the actual icon names to the combo box
                    icon_combobox.addItems(
                        [
                            "None",
                            "cross-platform",
                            "Android",
                            "cloud icon with a server",
                            "dollar (for paid asset)",
                            "iOS",
                            "Linux",
                            "Mac OS",
                            "paper note with a pen",
                            "USB key",
                            "Python script",
                            "general-purpose script",
                            "red question mark",
                            "Solaris",
                            "Web-related asset",
                            "Windows",
                        ]
                    )  # Replace this with your actual icon names
                    # Set the current index of the combo box to the data in the row
                    icon_combobox.setCurrentIndex(data[row][col])
                    # Connect the current index of the combo box to the handle_icon_change function
                    icon_combobox.currentIndexChanged.connect(
                        lambda index, row=row: self.handle_icon_change(
                            row, icon_combobox
                        )
                    )
                    # Set the cell widget of the row and column to the combo box
                    self.table.setCellWidget(row, col, icon_combobox)
                else:
                    # If the column is not the icon column
                    self.table.setItem(
                        row,
                        col,
                        QTableWidgetItem(data[row][col] if data[row][col] else ""),
                    )

    def resizeEvent(self, event):
        """
        DESCRIPTION: This function handles the resize event.
        param event:    The resize event
        return:    None
        """
        # Call the stretch_table method whenever the widget is resized

        self.stretch_table()

    def stretch_table(self):
        """
        DESCRIPTION: This function stretches the table to fit the window.
        return:    None
        """
        # self.adjustHeaders()

        total_width = self.table.width()
        num_columns = self.table.columnCount()

        fixed_columns = {
            5: 100  # Width of the icon column
        }

        # If column 0 has been resized, consider it fixed too
        if hasattr(self, 'is_column0_resized') and self.is_column0_resized:
            fixed_columns[0] = self.table.columnWidth(0)

        fixed_width_total = sum(fixed_columns.values())

        remaining_width = total_width - fixed_width_total
        stretchable_columns_count = num_columns - len(fixed_columns)

        # Set widths for fixed columns
        for col, width in fixed_columns.items():
            self.table.setColumnWidth(col, width)

        # If there are no stretchable columns, simply return
        if stretchable_columns_count == 0:
            return

        width_per_stretchable_column = int(remaining_width / stretchable_columns_count)

        # Distribute the remaining width among other columns
        for i in range(num_columns):
            if i not in fixed_columns:
                self.table.setColumnWidth(i, width_per_stretchable_column)
                remaining_width -= width_per_stretchable_column  # Adjust remaining width

                # Update width_per_stretchable_column for the next iterations
                stretchable_columns_count -= 1
                if stretchable_columns_count > 0:
                    width_per_stretchable_column = int(remaining_width / stretchable_columns_count)


class SelectionState(Enum):
    """
    DESCRIPTION: This class is an enumeration of the different selection states.
    """

    NONE = 1
    WORD_SELECTED = 2
    DESCRIPTION_SELECTION = 3
    DESCRIPTION_SELECTED = 4


class SplashScreenWindow(FramelessWindow):
    close_signal = Signal()  # Declare close_signal as a class attribute

    def __init__(self):
        super().__init__()
        # Create the main application window but don't show it yet
        self.mainWindow = MainWindow()

        self.resize(700, 600)
        # Correctly set the path for the splash screen image
        if getattr(sys, 'frozen', False):
            # The application is frozen
            bundle_dir = sys._MEIPASS
        else:
            # The application is not frozen
            bundle_dir = os.path.dirname(os.path.abspath(__file__))
        splash_image_path = os.path.join(bundle_dir, 'splashscreen.png')

        # Set window icon
        self.setWindowIcon(QIcon(splash_image_path))

        # Create splash screen and show window using CustomSplashScreen
        self.splashScreen = SplashScreen(QIcon(splash_image_path), self)
        self.splashScreen.setIconSize(QSize(150, 150))
        self.splashScreen.titleBar.hide()
        self.splashScreen.show()

        # Ensure GUI is updated
        QApplication.processEvents()

        # Connect the close_signal to the close_window method
        self.close_signal.connect(self.close_window)

        # Delay the rest of the initialization
        QTimer.singleShot(2000, self.createSubInterface)  # Delay for 3 seconds

    def createSubInterface(self):
        loop = QEventLoop(self)
        QTimer.singleShot(2000, loop.quit)  # Show splash screen for 3 seconds
        loop.exec()
        self.close_signal.emit()  # Emit the close_signal when createSubInterface is done

    def close_window(self):
        self.close()
        self.finishInit()  # Call finishInit when close_window is triggered

    def finishInit(self):
        # Close splash screen
        self.splashScreen.finish()

        # Show the main application window
        self.mainWindow.show()


class CustomCloseDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent, Qt.WindowTitleHint | Qt.CustomizeWindowHint)
        self.setWindowTitle("Save Changes")
        self.layout = QVBoxLayout(self)

        # Message label
        self.messageLabel = QLabel("Do you want to Save Table?", self)

        # Buttons
        self.yesButton = PushButton("Yes", self)
        self.noButton = PushButton("No", self)
        self.cancelButton = PushButton("Cancel", self)

        # Button layout
        buttonLayout = QHBoxLayout()
        buttonLayout.addWidget(self.yesButton)
        buttonLayout.addWidget(self.noButton)
        buttonLayout.addWidget(self.cancelButton)

        # Add widgets to layout
        self.layout.addWidget(self.messageLabel)
        self.layout.addLayout(buttonLayout)

        # Connect buttons to slots
        self.yesButton.clicked.connect(self.accept_yes)
        self.noButton.clicked.connect(self.accept_no)
        self.cancelButton.clicked.connect(self.reject)

        self.userChoice = None

        # Set fixed size for the dialog
        self.setFixedSize(self.sizeHint())

    def accept_yes(self):
        self.userChoice = QMessageBox.Yes
        self.accept()

    def accept_no(self):
        self.userChoice = QMessageBox.No
        self.accept()


class MainWindow(FluentWindow):
    def __init__(self):
        super().__init__()

        # Create sub interfaces
        self.combined = combine(self)

        self.initNavigation()
        self.initWindow()

        # Create a button for toggling the theme
        self.setMicaEffectEnabled(True)

    def toggleTheme(self):
        if self.current_theme == Theme.LIGHT:
            setTheme(Theme.DARK)
            self.current_theme = Theme.DARK
        else:
            setTheme(Theme.LIGHT)
            self.current_theme = Theme.LIGHT

        # Update the combo box style in PDFIndexCreator
        self.indexCreator.reapply_combo_box_style(self.current_theme)

    def initNavigation(self):
        # Add the sub interface with the appropriate icon and name
        self.addSubInterface(self.combined, FIF.DOCUMENT, 'Project')

        # add custom widget to bottom
        self.navigationInterface.addWidget(
            routeKey='avatar',
            widget= NavigationToolButton(FluentIcon.INFO),
            onClick=self.showMessageBox,
            position=NavigationItemPosition.BOTTOM,
        )

    def initWindow(self):
        self.resize(1900, 750)

        if getattr(sys, 'frozen', False):
            # The application is frozen
            bundle_dir = sys._MEIPASS
        else:
            # The application is not frozen
            bundle_dir = os.path.dirname(os.path.abspath(__file__))

        icon_path = os.path.join(bundle_dir, 'app_icon.png')
        self.setWindowIcon(QIcon(icon_path))

        self.setWindowTitle('PDF Index Creator')

        # Center the window on the screen
        desktop = QApplication.screens()[0].availableGeometry()
        w, h = desktop.width(), desktop.height()
        self.move(w // 2 - self.width() // 2, h // 2 - self.height() // 2)

    def closeEvent(self, event):
        """
        Handle the window close event.
        """
        if self.combined.is_saved:
            event.accept()
        else:
            # Create the custom dialog
            customDialog = CustomCloseDialog(self)
            self.combined.apply_dialog_stylesheet(customDialog)

            # Show the dialog and get the user's response
            customDialog.exec()

            if customDialog.userChoice == QMessageBox.Yes:
                if self.combined.save_table():
                    event.accept()
                else:
                    event.ignore()
            elif customDialog.userChoice == QMessageBox.No:
                event.accept()
            else:
                event.ignore()

    def showMessageBox(self):
        w = MessageBox(
            'Supporting the Author',
            'Personal development is not easy. If this project helps you, you can consider supporting me by buying me a coffee.',
            self
        )
        w.yesButton.setText('Sure!')
        w.cancelButton.setText('Maybe Next Time')

        if w.exec():
            QDesktopServices.openUrl(QUrl("https://github.com/kamalexandre/PDF-Index-Creator"))

class EventFilter(QObject):
    def __init__(self, message_box, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.message_box = message_box
        self.isDragging = False
        self.dragPosition = QPoint()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress:
            if event.button() == Qt.LeftButton:
                self.isDragging = True
                self.dragPosition = event.globalPosition().toPoint() - self.message_box.frameGeometry().topLeft()
                event.accept()
        elif event.type() == QEvent.MouseMove and self.isDragging:
            self.message_box.move(event.globalPosition().toPoint() - self.dragPosition)
            event.accept()
        elif event.type() == QEvent.MouseButtonRelease:
            self.isDragging = False
        return False  # Continue with the normal event processing


class CustomSplitter(QSplitter):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.allow_movement = True

    def mouseMoveEvent(self, event):
        if self.allow_movement:
            super().mouseMoveEvent(event)
        else:
            event.ignore()


class combine(QWidget):
    """Combine class for encapsulating the PDFIndexCreator and web view"""

    # Set the event filter to False
    event_filter_installed = False
    # Create a dictionary to store the buttons
    buttons = {}
    # Set the current row to None
    current_row = None

    # noinspection PyUnresolvedReferences
    def __init__(self, *args, **kwargs):
        """
        DESCRIPTION: This function initializes the main window.
        param args:    The arguments to pass to the parent class
        param kwargs:  The keyword arguments to pass to the parent class
        """
        super(combine, self).__init__(*args, **kwargs)
        self.setObjectName("Combined")
        self.last_op_undo_redo = False
        self.indexCreator = PDFIndexCreator(
            self
        )  # Pass the MainWindow instance into the PDFIndexCreator
        self.indexCreator.current_version = 0


        self.installAppEventFilter()
        self.page_number = None

        self.open_dialogs = []  # List to keep track of open search dialogs

        self.webview = FramelessWebEngineView(self)
        self.webview.settings().setAttribute(self.webview.settings().WebAttribute.PluginsEnabled, True)
        self.webview.settings().setAttribute(self.webview.settings().WebAttribute.PdfViewerEnabled, True)
        self.webview.settings().setAttribute(self.webview.settings().WebAttribute.WebGLEnabled, True)

        # Load the PDF.js viewer without specifying a PDF file
        self.webview.load(QUrl.fromUserInput(PDFJS))

        self.line_edit = LineEdit()
        self.line_edit.setReadOnly(False)
        self.line_edit.setPlaceholderText(
            "Word to be added as Index (also shows selected text)"
        )

        self.add_button = PrimaryPushButton('Add Word', self, FIF.ADD)
        self.add_button.setEnabled(True)
        self.add_button.clicked.connect(self.add_text)

        self.search_button = PrimaryPushButton('Search', self, FIF.SEARCH)
        self.search_button.setEnabled(True)
        self.search_button.clicked.connect(self.search_in_pdf)


        self.pre_page_edit = LineEdit()  # Add this
        self.pre_page_edit.setPlaceholderText(
            "Enter the number of pre-pages to Index Correctly"
        )
        self.pre_page_edit.setValidator(
            QIntValidator(0, 10000)
        )  # Add this (it allows only integer inputs from 0 to 10000)

        self.book_edit = LineEdit()  # Add this
        self.book_edit.setPlaceholderText("Enter Book number if needed for idx file")
        self.book_edit.setValidator(
            QIntValidator(0, 10000)
        )  # Add this (it allows only integer inputs from 0 to 10000)

        self.generate_idx_button = PushButton("Generate IDX file")
        self.generate_idx_button.setEnabled(True)
        self.generate_idx_button.clicked.connect(self.generate_idx)

        self.splitter = CustomSplitter(Qt.Orientation.Horizontal)
        self.splitter.addWidget(self.webview)
        self.splitter.addWidget(self.indexCreator)
        self.splitter.setStretchFactor(0, 1)  # Set stretch factor for the first widget (self.webview)
        self.splitter.setStretchFactor(1, 1)  # Set stretch factor for the second widget (self.indexCreator)
        self.splitter.splitterMoved.connect(self.handle_splitter_moved)
        self.splitter_moved_manually = False  # Add this attribute to keep track of manual splitter movement
        self.splitter.setStyleSheet("""
                QSplitter::handle {
                    background-color: transparent;
                }
                QSplitter::handle:horizontal {
                    width: 5px;
                }
                QSplitter::handle:vertical {
                    height: 5px;
                }
            """)



        MenuLayout = QHBoxLayout()
        self.commandBar = CommandBar(self)
        MenuLayout.addWidget(self.commandBar, 0)
        self.commandBar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.commandBarButton(FluentIcon.LIBRARY, 'New Project')
        self.commandBar.addSeparator()
        self.commandBarButton(FluentIcon.DICTIONARY_ADD, 'Open PDF')
        self.commandBarButton(FluentIcon.LAYOUT, 'Import Table')
        self.commandBar.addSeparator()
        self.commandBarButton(FluentIcon.SAVE, 'Save Table')
        self.commandBarButton(FluentIcon.SAVE_AS, 'Save Table As')

        self.commandBar.addSeparator()
        self.pdfVisibilityAction = Action(FluentIcon.HIDE, 'Hide PDF Viewer', checkable=True)
        self.pdfVisibilityAction.triggered.connect(self.toggle_splitter_position)
        self.commandBar.addAction(self.pdfVisibilityAction)
        self.pdfVisibilityAction.setChecked(True)  # Initially visible

        # Create a button for toggling the theme
        self.commandBar.addSeparator()
        self.themeToggleAction = Action(FluentIcon.CONSTRACT, '')  # Assuming FluentIcon.THEME exists
        self.themeToggleAction.triggered.connect(self.toggleTheme)
        self.commandBar.addAction(self.themeToggleAction)

        self.current_theme = Theme.LIGHT

        self.search_tables = []

        # Initialize file_path to None
        self.file_path = None

        # Initialize update counter and auto-save threshold
        self.update_counter = 0
        self.auto_save_threshold = 50
        self.indexCreator.table.itemChanged.connect(self.handle_item_changed)

        # Initialize save flag
        self.is_saved = False

        # Initialize system tray icon and menu
        self.tray_icon = None
        self.tray_menu = None
        self.closing_program = False
        # Create status bar

        layout = QVBoxLayout()
        layout.addLayout(MenuLayout)
        layout.addWidget(self.splitter)

        h_layout = QHBoxLayout()
        h_layout.addWidget(self.line_edit)
        h_layout.addWidget(self.add_button)
        h_layout.addWidget(self.search_button)
        h_layout.addWidget(self.pre_page_edit)  # Add this
        h_layout.addWidget(self.book_edit)
        h_layout.addWidget(self.generate_idx_button)
        # Add stretch to the layout to distribute the available space

        layout.addLayout(h_layout)

        self.selection_timer = QTimer(self)
        self.selection_timer.setSingleShot(True)
        self.selection_timer.timeout.connect(self.selection_changed)
        self.webview.selectionChanged.connect(self.selection_changed)

        self.mouse_released = False
        self.button = None

        self.setLayout(layout)  # Set the layout for the combine class
        self.toggle_splitter_position()

    def toggleTheme(self):
        if self.current_theme == Theme.LIGHT:
            setTheme(Theme.DARK)
            self.current_theme = Theme.DARK
        else:
            setTheme(Theme.LIGHT)
            self.current_theme = Theme.LIGHT

        # Update the combo box style in PDFIndexCreator
        self.indexCreator.reapply_combo_box_style(self.current_theme)

        # Apply the new theme to all open search dialogs
        for dialog in self.open_dialogs:
            self.apply_dialog_stylesheet(dialog)

    def commandBarButton(self, icon, text):
        # Define a dictionary to map the text labels to their corresponding methods.
        action_methods = {
            'New Project': self.new_file,
            'Open PDF': self.open_pdf_file,
            'Save Table': self.save_table,
            'Save Table As': self.save_table_as_excel,
            'Import Table': self.open_table_from_excel
        }

        # Create an action object.
        action = Action(icon, text, self)

        # If the text label exists in the dictionary, connect its corresponding method.
        # Otherwise, just print the text.
        if text in action_methods:
            action.triggered.connect(action_methods[text])
        else:
            action.triggered.connect(lambda: print(text))

        # Add the action to the command bar.
        self.commandBar.addAction(action)

    def handle_splitter_moved(self, pos, index):
        # Update the splitter_moved_manually attribute whenever the splitter is moved
        self.splitter_moved_manually = True

        # Check if the splitter has been moved to a position of 0
        if self.splitter.sizes()[0] == 0:
            self.pdfVisibilityAction.setText('Show PDF Viewer')
            self.pdfVisibilityAction.setIcon(FluentIcon.VIEW)  # Assuming you have a FluentIcon.VIEW
            self.pdfVisibilityAction.setChecked(False)  # Uncheck the button
            self.splitter_moved_manually = False  # Reset the flag
        elif self.splitter.sizes()[0] > 0:
            self.pdfVisibilityAction.setText('Hide PDF Viewer')
            self.pdfVisibilityAction.setIcon(FluentIcon.HIDE)  # Assuming you have a FluentIcon.HIDE
            self.pdfVisibilityAction.setChecked(True)  # Uncheck the button
            self.splitter_moved_manually = True  # Reset the flag

    def toggle_splitter_position(self):
        current_position = self.splitter.sizes()[0]
        current_text = self.pdfVisibilityAction.text()

        for key in list(self.buttons.keys()):  # make a copy of the keys
            self.buttons[key].deleteLater()
            self.buttons.pop(key)

        if current_position > 0:
            # print("1")
            self.splitter.setSizes([0, self.splitter.width()])
            self.pdfVisibilityAction.setText('Show PDF Viewer')
            self.pdfVisibilityAction.setIcon(FluentIcon.VIEW)  # Assuming you have a FluentIcon.VIEW
            # self.webview.setMaximumWidth(0)  # Prevent the webview from expanding
            # self.indexCreator.setMaximumWidth(self.splitter.width())  # Set the maximum width of indexCreator to the width of the splitter
            self.splitter_moved_manually = False  # Reset the flag
        elif self.splitter_moved_manually and current_position == 0:
            # print("2")
            self.pdfVisibilityAction.setText('Show PDF Viewer')
            self.pdfVisibilityAction.setIcon(FluentIcon.VIEW)  # Assuming you have a FluentIcon.VIEW
            # pass
            self.splitter_moved_manually = False  # Reset the flag
        elif not self.splitter_moved_manually:
            # print("3")
            half_width = self.splitter.width() / 2
            self.splitter.setSizes([half_width, half_width])
            self.pdfVisibilityAction.setText('Hide PDF Viewer')
            self.pdfVisibilityAction.setIcon(FluentIcon.HIDE)  # Assuming you have a FluentIcon.HIDE
            # self.webview.setMaximumWidth(
            #    16777215)  # Reset the maximum width of webview (16777215 is the default maximum width)
            # self.indexCreator.setMaximumWidth(16777215)  # Reset the maximum width of indexCreator
            self.splitter_moved_manually = False  # Reset the flag

    def reset_to_default_state(self):
        """
        DESCRIPTION: This function resets the window to the default state.
        return: None
        """
        # Now, permanently delete all buttons
        for key in list(self.buttons.keys()):
            self.buttons[key].deleteLater()
            del self.buttons[key]
        # Window setup


        # Line edits and other input setup
        self.line_edit.clear()
        self.pre_page_edit.clear()
        self.book_edit.clear()

        # Reset the web view

        PDF = None
        self.webview.load(QUrl.fromUserInput("%s?file=%s" % (PDFJS, PDF)))

        # Reset internal states
        self.selection_state = SelectionState.NONE
        self.file_path = None
        self.update_counter = 0
        self.is_saved = False

        # Reset the PDFIndexCreator state
        self.indexCreator.history = []
        self.indexCreator.current_version = -1
        self.indexCreator.clear_table()  # Assuming this method clears the table

        self.indexCreator.handle_icon_column_visibility(False)
        self.indexCreator.handle_book_column_visibility(False)
        self.indexCreator.handle_reference_column_visibility(False)

        ### Restore the default state of the command bar
        [action.setChecked(False) for action in self.indexCreator.commandBar.actions() if action.isCheckable()]

        # Clear status message
        # self.status_bar.clearMessage()

        # Reset splitter to the middle
        splitter_size = self.splitter.size()
        half_size = splitter_size.width() // 2
        self.splitter.setSizes([half_size, half_size])
        self.pdfVisibilityAction.setText('Hide PDF Viewer')
        self.pdfVisibilityAction.setIcon(FluentIcon.HIDE)  # Assuming you have a FluentIcon.HIDE
        self.pdfVisibilityAction.setChecked(True)  # Uncheck the button
        self.splitter_moved_manually = False  # Reset the flag

    def generate_idx(self):
        """
        DESCRIPTION: This function generates the idx file.
        param:         None
        return:    None
        """

        # Create a dictionary of icon names to their corresponding file names
        icon_mapping = {
            "cross-platform": "\\all",
            "Android": "\\android",
            "cloud icon with a server": "\\cloud",
            "dollar (for paid asset)": "\\coin",
            "iOS": "\\ios",
            "Linux": "\\linux",
            "Mac OS": "\\mac",
            "paper note with a pen": "\\note",
            "USB key": "\\portable",
            "Python script": "\\python",
            "general-purpose script": "\\script",
            "red question mark": "\\question",
            "Solaris": "\\solaris",
            "Web-related asset": "\\web",
            "Windows": "\\win",
        }

        # Dictionary to store entries for each book by book number
        book_entries_dict = {}
        no_book_entries = []  # List to store entries with no book number

        for row in range(self.indexCreator.table.rowCount()):
            # Get the word and page number from the current row
            word = self.indexCreator.table.item(row, 0).text()
            # Check if the word is empty or None
            if not word.strip():
                continue  # Skip this row and move to the next one

            page = self.indexCreator.table.item(row, 1).text()
            book_number = self.indexCreator.table.item(row, 3).text()
            references = (
                self.indexCreator.table.item(row, 4).text().split("\n")
            )  # Split by newlines

            # Get the combo box widget for the "Icon" column in the current row
            icon_combobox = self.indexCreator.table.cellWidget(row, 5)
            icon_text = (
                icon_combobox.currentText()
                if isinstance(icon_combobox, QComboBox)
                else "None"
            )

            icon = ""  # Default to empty string
            if icon_text != "None":  # If an icon is selected
                icon = icon_mapping.get(icon_text, "")  # Translate the icon text to LaTeX command

            # Start creating the index entry with the word.
            # print(word)
            index_entry = f"\\indexentry{{{word}"

            # If reference exists and are not empty or white space, add them after the word.
            if any(reference.strip() for reference in references):
                # Add the icon after the reference if it exists.
                index_entry += (
                    f"!{' '.join(reference for reference in references if reference.strip())} {icon}"
                    if icon
                    else f"!{' '.join(reference for reference in references if reference.strip())}"
                )
            elif icon:  # Add only the icon if there are no references
                index_entry += f" {icon}"

            # If book number exists, format with icon and book number.
            if book_number.strip():
                if icon:
                    # The icon is placed before "|book{...}" command with a preceding space
                    index_entry += f" {icon}|book{{{book_number}}}"
                else:
                    # If no icon, just add the book number
                    index_entry += f"|book{{{book_number}}}"
            else:
                # If there's no book number but an icon exists, add it directly after the word/reference
                index_entry += f"{icon}" if icon else ""

            # Finalize the index entry by adding the page number.
            if "-" in page:  # Check if the page is a range
                page_start, page_end = page.split('-')
                page_range = page if page_start != page_end else page_start  # Only use range if start != end
            else:
                page_number = int(page)  # Convert to an integer to avoid leading zeros
                page_range = f"{page_number}"  # Single page, no range

            index_entry += f"}}{{{page_range}}}"

            # Check if the book_number is already in the dictionary
            if not book_number.strip():
                no_book_entries.append(index_entry)
            else:
                if book_number not in book_entries_dict:
                    # Create a new section comment for the book
                    section_comment = (
                        f"\\comment{{********************************************************************************************\n"
                        f"                                       BOOK {book_number} SECTION\n"
                        f"*****************************************************************************************************}}"
                    )
                    book_entries_dict[book_number] = [section_comment]
                # Append the index entry to the correct book section
                book_entries_dict[book_number].append(index_entry)

        # Add section comment for entries with no book number
        if no_book_entries:
            section_comment = (
                f"\\comment{{********************************************************************************************\n"
                f"                                       NO BOOK SECTION\n"
                f"*****************************************************************************************************}}"
            )
            book_entries_dict['none'] = [section_comment] + no_book_entries

        # Sort the dictionary by book number and join the index entries for each book into a single string
        sorted_books = sorted(book_entries_dict.keys(), key=lambda x: (x != 'none', x))
        index_string = "\n".join("\n".join(book_entries_dict[book]) for book in sorted_books)

        # Prompt the user to select the file path and name
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save IDX File", "", "IDX Files (*.idx);;All Files (*.*)"
        )

        if file_path:
            try:
                with open(file_path, "w") as file:
                    file.write(index_string)
                print(f"Index entries saved to: {file_path}")
            except IOError:
                print("Error saving IDX file.")

    def add_text(self):
        """
        DESCRIPTION: This function adds the selected text to the table.
        return:        None
        """
        book_number = self.book_edit.text().strip()
        selected_text = self.line_edit.text().strip()
        if selected_text:
            self.indexCreator.table.clearSelection()
            row_position = self.indexCreator.table.rowCount()
            self.indexCreator.table.insertRow(row_position)
            self.indexCreator.table.setItem(
                row_position, 0, QTableWidgetItem(selected_text)
            )
            self.indexCreator.table.setItem(
                row_position, 1, QTableWidgetItem(str(self.page_number))
            )
            self.indexCreator.table.setItem(row_position, 2, QTableWidgetItem(" "))
            if book_number:
                self.indexCreator.table.setItem(
                    row_position, 3, QTableWidgetItem(book_number)
                )
            else:
                self.indexCreator.table.setItem(row_position, 3, QTableWidgetItem(" "))
            self.indexCreator.table.setItem(row_position, 4, QTableWidgetItem(" "))
            self.indexCreator.table.setItem(row_position, 5, QTableWidgetItem("None"))

            # Save the current row
            self.current_row = row_position

            if self.button is not None:
                self.button.deleteLater()
                self.button = None

            self.selection_state = SelectionState.WORD_SELECTED
            self.line_edit.clear()

            # Update the current row to the last row
            self.current_row = row_position

            for key in list(self.buttons.keys()):  # make a copy of the keys
                self.buttons[key].deleteLater()
                self.buttons.pop(key)

            self.indexCreator.table.clearSelection()

            # Set the selection state to DESCRIPTION_SELECTION
            self.selection_state = SelectionState.DESCRIPTION_SELECTION

            # Populate the "Icon" column for the newly added row
            self.indexCreator.populate_icon_column(row_position)

            # Save the state after a word is added
            self.indexCreator.save_state()

    def open_table_from_excel(self):
        """
        DESCRIPTION: This function opens a table from an Excel file.
        return:    None
        """
        # Temporarily hide all buttons
        self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")
        for button in self.buttons.values():
            button.hide()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Table from Excel", "", "Excel Files (*.xlsx);;All Files (*.*)"
        )
        if file_path:
            self.load_table_from_excel(file_path)

    def load_table_from_excel(self, file_path):
        """
        DESCRIPTION: This function loads a table from an Excel file.
        param file_path:           The path of the Excel file.
        return:            None
        """
        self.indexCreator.handle_icon_column_visibility(False)
        self.indexCreator.handle_book_column_visibility(False)
        self.indexCreator.handle_reference_column_visibility(False)
        # self.indexCreator.commandBar.actions()[5].setChecked(False)
        ### Restore the default state of the command bar
        [action.setChecked(False) for action in self.indexCreator.commandBar.actions() if action.isCheckable()]
        self.indexCreator.future_states = []
        self.indexCreator.history = []
        self.indexCreator.current_version = 0
        self.update_counter = 0
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        rows = sheet.max_row
        columns = sheet.max_column

        self.indexCreator.table.clear()
        self.indexCreator.table.setRowCount(rows - 1)
        self.indexCreator.table.setColumnCount(columns)

        # Populate header row
        for column in range(columns):
            header = sheet.cell(row=1, column=column + 1).value
            self.indexCreator.table.setHorizontalHeaderItem(
                column, QTableWidgetItem(header)
            )
            self.update_counter = 0

        # Populate data rows
        for row in range(2, rows + 1):
            for column in range(1, columns + 1):
                value = sheet.cell(row=row, column=column).value

                if column != 6:  # If it's not the "Icon" column
                    self.indexCreator.table.setItem(
                        row - 2,
                        column - 1,
                        QTableWidgetItem(str(value) if value is not None else ""),
                    )
                else:  # If it's the "Icon" column
                    self.indexCreator.populate_icon_column(
                        row - 2
                    )  # Populate Icon column with default value
                    if value is not None:
                        self.indexCreator.table.setItem(
                            row - 2, column - 1, QTableWidgetItem(str(value))
                        )
            self.update_counter = 0
        workbook.close()

        # Update ComboBox indices after all items have been created
        for row in range(2, rows + 1):
            value = sheet.cell(row=row, column=6).value  # Icon column's value
            if value is not None:
                combobox = self.indexCreator.table.cellWidget(
                    row - 2, 5
                )  # Get the combobox

                index = combobox.findText(str(value))  # Find the index of the item
                if index != -1:  # If the item was found
                    combobox.setCurrentIndex(index)  # Set the current index
            self.update_counter = 0
        # Update file_path and is_saved flag
        self.file_path = file_path
        self.is_saved = True

        self.indexCreator.history = []
        self.indexCreator.current_version = -1
        self.update_counter = 0
        self.indexCreator.future_states = []
        self.indexCreator.history = []
        self.indexCreator.current_version = 0
        self.indexCreator.save_state()
        # Notify the user about the successful load
        print(f"Table loaded from Excel: {file_path}")

    def create_tray_icon(self):
        """
        DESCRIPTION: This function creates a tray icon.
        return:        None
        """
        self.tray_icon = QSystemTrayIcon(self)
        self.tray_icon.setIcon(
            self.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxInformation)
        )
        self.tray_icon.setToolTip("PDF Index Creator")
        self.tray_icon.show()
        self.update_counter = 0

    def show_auto_save_notification(self):
        """
        DESCRIPTION: This function shows a notification when the table is auto-saved.
        return:        None
        """
        InfoBar.success(
            title='Success',  # Adjusted the title from 'Lesson 4' to match the previous notification's title
            content="Table Saved.",  # Adjusted the content to match the previous notification's message
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP_RIGHT,
            # position='Custom',   # NOTE: use custom info bar manager
            duration=2000,
            parent=self
        )
        self.update_counter = 0

    def show_auto_save_notification_error(self):
        InfoBar.error(
            title='Nothing to Save',  # Adjusted the title from 'Lesson 4' to match the previous notification's title
            content="Table Is Empty.",  # Adjusted the content to match the previous notification's message
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP_RIGHT,
            # position='Custom',   # NOTE: use custom info bar manager
            duration=2000,
            parent=self
        )
        self.update_counter = 0

    def open_pdf_file(self):
        """
        DESCRIPTION: This function opens a PDF file.
        return:    None
        """
        self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")
        # Temporarily hide all buttons
        for button in self.buttons.values():
            button.hide()
        pdf_file_path, _ = QFileDialog.getOpenFileName(
            self, "Open PDF", "", "PDF Files (*.pdf);;All Files (*.*)"
        )
        # print(pdf_file_path)
        if pdf_file_path:
            print(pdf_file_path)
            self.load_pdf(pdf_file_path)

    def load_pdf(self, pdf_file_path):
        """
        DESCRIPTION: This function loads a PDF file.
        param pdf_file_path:    The path of the PDF file.
        return:        None
        """
        # Reset splitter to the middle
        splitter_size = self.splitter.size()
        half_size = splitter_size.width() // 2
        self.splitter.setSizes([half_size, half_size])
        self.pdfVisibilityAction.setText('Hide PDF Viewer')
        self.pdfVisibilityAction.setIcon(FluentIcon.HIDE)  # Assuming you have a FluentIcon.HIDE
        self.pdfVisibilityAction.setChecked(True)  # Uncheck the button
        self.splitter_moved_manually = False  # Reset the flag
        # self.webview.load(QUrl.fromUserInput(file_path))
        # PDF_PATH = "http://127.0.0.1:8000/pdf?file=" + quote(pdf_file_path)
        # PDF_URL = "http://127.0.0.1:8000/pdfjs/web/viewer.html?file=" + quote(PDF_PATH)
        # self.webview.load(QUrl(PDF_URL))

        # PDF = None
        self.webview.load(QUrl.fromUserInput("%s?file=%s" % (PDFJS, pdf_file_path)))
        # self.webview.load(QUrl.fromUserInput("%s?file=%s" % (PDFJS, pdf_file_path)))
        # self.update_counter = 0  # Reset the update counter
        # self.is_saved = False  # Reset the save flag
        # self.file_path = None  # Clear the current file path
        self.current_pdf_path = pdf_file_path  # Store the path for later use in search
        print(f"PDF loaded: {pdf_file_path}")
        self.selection_state = SelectionState.NONE



    def new_file(self):
        for key in list(self.buttons.keys()):
            self.buttons[key].deleteLater()
            self.buttons.pop(key)

        # If the content is saved, or the table is empty, reset to default without prompting
        if self.is_saved or self.is_table_empty():
            self.reset_to_default_state()
        else:
            # Create the custom dialog
            customDialog = CustomCloseDialog(self)
            self.apply_dialog_stylesheet(customDialog)  # Assuming this method styles the dialog

            # Show the dialog and get the user's response
            customDialog.exec()

            if customDialog.userChoice == QMessageBox.Yes:
                saved = self.save_table()
                if saved:
                    self.reset_to_default_state()
            elif customDialog.userChoice == QMessageBox.No:
                self.reset_to_default_state()

    def save_table(self):
        """
        DESCRIPTION: This function saves the table.
        return: True if the table was saved successfully, False otherwise.
        """
        # Check if table is empty

        if self.file_path is None:
            if self.is_table_empty():
                # print("empty")
                self.show_auto_save_notification_error()
                return False
            else:
                return self.save_table_as_excel()
        else:
            self.save_data_to_excel(self.file_path)
            return True

    def is_table_empty(self):
        """
        DESCRIPTION: This function checks if the table is empty.
        return: True if the table is empty, False otherwise.
        """
        for row in range(self.indexCreator.table.rowCount()):
            for column in range(self.indexCreator.table.columnCount()):
                item = self.indexCreator.table.item(row, column)
                if item is not None and item.text():
                    return False  # A non-empty item was found, table is not empty
        return True  # All items checked were empty, table is empty

    def save_table_as_excel(self):
        """
        DESCRIPTION: This function saves the table as an Excel file.
        return: True if the table was saved successfully, False otherwise.
        """
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Table as Excel", "", "Excel Files (*.xlsx);;All Files (*.*)"
        )
        if file_path:
            self.save_data_to_excel(file_path)
            self.file_path = file_path
            self.is_saved = True
            self.show_auto_save_notification()
            return True
        return False

    def save_data_to_excel(self, file_path):
        """
        DESCRIPTION: This function saves the data to an Excel file.
        param file_path:           The path of the Excel file.
        return:            None
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Define the alignment setting (wrap text and top alignment)
        alignment_settings = openpyxl.styles.Alignment(wrapText=True, vertical='top')
        wrap_alignment = openpyxl.styles.Alignment(wrapText=True)

        # Write header row with only wrap text
        for column in range(self.indexCreator.table.columnCount()):
            header = self.indexCreator.table.horizontalHeaderItem(column).text()
            cell = sheet.cell(row=1, column=column + 1, value=header)
            cell.alignment = wrap_alignment

        # Write data rows with both wrap text and top alignment
        for row in range(self.indexCreator.table.rowCount()):
            for column in range(self.indexCreator.table.columnCount()):
                item = self.indexCreator.table.item(row, column)
                if item is not None:
                    value = item.text()
                    cell = sheet.cell(row=row + 2, column=column + 1, value=value)
                    cell.alignment = alignment_settings

        # Adjust column width to fit data
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # additional +2 for some padding
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

        workbook.save(file_path)
        workbook.close()

        # Reset update counter after successful save

        # Set save flag to True
        if not self.is_saved:
            pass
        if self.is_saved:
            # Show auto-save notification
            self.show_auto_save_notification()

        # Notify the user about the successful save
        print(f"Table saved as Excel: {file_path}")

    def closeEvent(self, event):
        """
        DESCRIPTION: This function closes the program.
        param event:               The event.
        return:        None
        """
        # if event is None or not self.closing_program:
        #    server.stop_server()
        #    if event is not None:
        #        event.accept()
        #    return

        if self.is_saved:
            # server.stop_server()
            event.accept()
        else:
            reply = QMessageBox.question(
                self,
                "Save Changes",
                "Do you want to save before closing?",
                QMessageBox.StandardButton.Yes
                | QMessageBox.StandardButton.No
                | QMessageBox.StandardButton.Cancel,
            )

            if reply == QMessageBox.StandardButton.Yes:
                self.save_table()
                # server.stop_server()
                event.accept()
            elif reply == QMessageBox.StandardButton.No:
                # server.stop_server()
                event.accept()
            else:
                event.ignore()

    def handle_item_changed(self):
        """
        DESCRIPTION: This function handles the item changed event.
        return:        None
        """
        # Increment the update counter
        # print(self.update_counter)
        if not self.last_op_undo_redo:
            if self.is_saved:
                self.update_counter += 0.5

            # Check if the auto-save threshold is reached
            if self.update_counter >= self.auto_save_threshold:
                self.save_table()

        # else:
        # print("here not stretch table")
        # print("here to stretch table")
        self.indexCreator.adjustHeaders()
        # self.indexCreator.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.indexCreator.stretch_table()

    def on_word_button_clicked(self):
        """
        DESCRIPTION: This function handles the word button clicked event.
        return:    None
        """
        # selected_text = self.line_edit.text().strip()
        selected_text = re.sub(r'\s+', ' ', self.line_edit.text().replace('\n', ' ')).strip()

        book_number = self.book_edit.text().strip()

        # Check if a row is selected in the table
        if self.indexCreator.table.selectionModel().hasSelection():
            current_row = self.indexCreator.table.selectionModel().currentIndex().row()
        else:
            current_row = self.indexCreator.table.rowCount()
            self.indexCreator.table.insertRow(current_row)

        # Populate the table with the selected data
        self.indexCreator.table.setItem(
            current_row, 0, QTableWidgetItem(selected_text)
        )
        self.indexCreator.table.setItem(
            current_row, 1, QTableWidgetItem(str(self.page_number))
        )

        # Only update the "Book" column if a new row is added
        if not self.indexCreator.table.selectionModel().hasSelection():
            self.indexCreator.table.setItem(current_row, 2, QTableWidgetItem(" "))
            if book_number:
                self.indexCreator.table.setItem(
                    current_row, 3, QTableWidgetItem(book_number)
                )
            else:
                self.indexCreator.table.setItem(current_row, 3, QTableWidgetItem(" "))
            self.indexCreator.table.setItem(current_row, 4, QTableWidgetItem(" "))
            self.indexCreator.table.setItem(current_row, 5, QTableWidgetItem("None"))

        # Save the current row
        self.current_row = current_row

        if self.button is not None:
            self.button.deleteLater()
            self.button = None

        self.selection_state = SelectionState.WORD_SELECTED
        self.line_edit.clear()

        # Clear the buttons (if any exist)
        for key in list(self.buttons.keys()):
            self.buttons[key].deleteLater()
            self.buttons.pop(key)

        self.indexCreator.table.clearSelection()
        self.selection_state = SelectionState.DESCRIPTION_SELECTION

        self.indexCreator.populate_icon_column(current_row)
        self.indexCreator.save_state()
        # Clear the selection in the QWebEngineView.
        self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")

    def on_add_desc_button_clicked(self):
        """
        DESCRIPTION: This function handles the add description button clicked event.
        return:    None
        """
        # selected_text = self.line_edit.text().strip()
        # selected_text = self.line_edit.text().strip().replace('\n', ' ')
        selected_text = re.sub(r'\s+', ' ', self.line_edit.text().replace('\n', ' ')).strip()
        if self.current_row is None:
            return

        if self.indexCreator.table.selectionModel().hasSelection():
            current_row = self.indexCreator.table.selectionModel().currentIndex().row()
        else:
            current_row = self.indexCreator.table.rowCount() - 1

        current_text = self.indexCreator.table.item(current_row, 2).text()

        if current_text:
            if current_text == " ":
                new_text = "- " + selected_text
            else:
                new_text = current_text + "\n- " + selected_text
        else:
            new_text = "- " + selected_text

        self.indexCreator.table.setItem(current_row, 2, QTableWidgetItem(new_text))

        if self.button is not None:
            self.button.deleteLater()
            self.button = None

        self.create_button(
            "Add Description",
            self.on_add_desc_button_clicked,
            "addDescriptionButton",
            100,
        )
        self.selection_state = SelectionState.DESCRIPTION_SELECTED
        for key in list(self.buttons.keys()):  # make a copy of the keys
            self.buttons[key].deleteLater()
            self.buttons.pop(key)
        # self.indexCreator.table.clearSelection()

        # self.indexCreator.table.clearSelection()  # Clear selection

        # Save the state after a description is added
        self.indexCreator.save_state()
        self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")

    def on_add_ref_button_clicked(self):
        """
        DESCRIPTION: This function handles the add reference button clicked event.
        return:    None
        """
        # selected_text = self.line_edit.text().strip()
        selected_text = re.sub(r'\s+', ' ', self.line_edit.text().replace('\n', ' ')).strip()
        if self.current_row is None:
            return

        if self.indexCreator.table.selectionModel().hasSelection():
            current_row = self.indexCreator.table.selectionModel().currentIndex().row()
        else:
            current_row = self.indexCreator.table.rowCount() - 1

        current_text = self.indexCreator.table.item(current_row, 4).text()

        if current_text:
            if current_text == " ":
                new_text = selected_text
            else:
                new_text = current_text + "\n" + selected_text
        else:
            new_text = selected_text

        self.indexCreator.table.setItem(current_row, 4, QTableWidgetItem(new_text))

        if self.button is not None:
            self.button.deleteLater()
            self.button = None

        # self.create_button("Add Description", self.on_add_desc_button_clicked, "addDescriptionButton", 100)
        self.selection_state = SelectionState.DESCRIPTION_SELECTED
        for key in list(self.buttons.keys()):  # make a copy of the keys
            self.buttons[key].deleteLater()
            self.buttons.pop(key)
        # self.indexCreator.table.clearSelection()

        # self.indexCreator.table.clearSelection()  # Clear selection

        # Save the state after a description is added
        self.indexCreator.save_state()
        self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")

    def eventFilter(self, obj, event):
        """
        DESCRIPTION: This function handles the event filter.
        param obj:     The object that is being filtered.
        param event:   The event that is being filtered.
        return:    None
        """
        if event.type() == QEvent.Type.MouseButtonPress and event.button() == Qt.MouseButton.LeftButton:
            self.get_current_page_number()

        if event.type() == QEvent.Type.MouseButtonRelease and event.button() == Qt.MouseButton.LeftButton:
            self.selection_timer.stop()

        return super().eventFilter(obj, event)

    def installAppEventFilter(self, **kwargs):
        """
        DESCRIPTION: This function installs the event filter for page number extraction.
        return:    None
        :param **kwargs:
        """
        if not self.event_filter_installed:
            QApplication.instance().installEventFilter(self)
            self.event_filter_installed = True

    def selection_changed(self):
        """
        DESCRIPTION: This function handles the selection changed event.
        return:    None
        """
        self.selection_timer.start(100)  # waits for 100 milliseconds before handling the selection
        table_rows_count = self.indexCreator.table.rowCount()
        # print(table_rows_count)
        if table_rows_count == 0:
            self.selection_state = SelectionState.NONE

        selected_text = self.webview.selectedText()
        if selected_text:
            self.line_edit.setText(selected_text)
            if self.selection_state in [
                SelectionState.NONE,
                SelectionState.WORD_SELECTED,
            ]:
                self.create_button(
                    "Add Word", self.on_word_button_clicked, "addWordButton", 0
                )
                self.selection_state = SelectionState.WORD_SELECTED
            if self.selection_state in [
                SelectionState.DESCRIPTION_SELECTION,
                SelectionState.DESCRIPTION_SELECTED,
            ]:
                self.create_button(
                    "Add Word", self.on_word_button_clicked, "addWordButton", 0
                )
                self.create_button(
                    "Add Description",
                    self.on_add_desc_button_clicked,
                    "addDescriptionButton",
                    100,
                )
                self.create_button(
                    "Add Reference",
                    self.on_add_ref_button_clicked,
                    "addReferenceButton",
                    100,
                )
                self.selection_state = SelectionState.DESCRIPTION_SELECTED
        else:
            self.line_edit.clear()
            for key in list(self.buttons.keys()):
                self.buttons[key].deleteLater()
                self.buttons.pop(key)
            if self.selection_state in [
                SelectionState.WORD_SELECTED,
                SelectionState.DESCRIPTION_SELECTED,
            ]:
                self.selection_state = SelectionState.DESCRIPTION_SELECTED

    def create_button(self, text, callback, button_key, offset):
        """
        DESCRIPTION: This function creates a button.
        param text:        The text of the button.
        param callback:    The callback function.
        param button_key:  The key of the button.
        param offset:      The offset of the button.
        return:    None
        """
        # If button already exists, delete and remove from dict
        if button_key in self.buttons:
            self.buttons[button_key].deleteLater()
            del self.buttons[button_key]

        button = QPushButton(text, self)
        button.setObjectName(button_key)

        # Define default color
        color = "#7BDCB5"  # green in hexadecimal

        if button_key == "addReferenceButton":
            color = "#F39A2B"  # red in hexadecimal

        if button_key == "addDescriptionButton":
            color = "#4773aa"  # blue in hexadecimal

        if button_key == "addWordButton":
            color = "#7cb287"  # green in hexadecimal

        # Calculate a darker tone for hover effect
        hover_color = "#{:02X}{:02X}{:02X}".format(
            max(0, int(color[1:3], 16) - 20),
            max(0, int(color[3:5], 16) - 20),
            max(0, int(color[5:7], 16) - 20)
        )

        # Apply the shadow effect
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(10)  # Adjust for desired "blur" effect
        shadow.setColor(Qt.black)
        shadow.setOffset(0)
        button.setGraphicsEffect(shadow)

        # The button's style with simulated before and after circles
        button.setStyleSheet(
            f"""
            QPushButton {{
                border-radius: 10px;
                background-color: {color};
                color: white;
                border-image: radial-gradient(circle at 10% 10%, transparent, transparent 20%, {color} 21%, {color} 79%, transparent 80%, transparent 90%, {color});
            }}
            QPushButton:hover {{
                background-color: {hover_color};
            }}
            QPushButton:pressed {{
                background-color: #555555;
            }}
            """
        )

        button.setFlat(True)
        button.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))

        # Calculate the position based on cursor
        global_cursor_position = QCursor.pos()
        local_cursor_position = self.webview.mapFromGlobal(global_cursor_position)

        # Check for reverse selection
        if not hasattr(self, 'last_cursor_position'):
            self.last_cursor_position = local_cursor_position

        reverse_horizontal_selection = local_cursor_position.x() < self.last_cursor_position.x()
        reverse_vertical_selection = local_cursor_position.y() < self.last_cursor_position.y()

        self.last_cursor_position = local_cursor_position

        # Vertical adjustment
        vertical_offset = 20  # Adjust this value based on your preference
        if reverse_vertical_selection:
            button_y = local_cursor_position.y() - vertical_offset
        else:
            button_y = local_cursor_position.y() + vertical_offset

        # Horizontal adjustment
        horizontal_offset = 5  # Adjust this value based on your preference
        if reverse_horizontal_selection:
            button_x = local_cursor_position.x() - horizontal_offset
        else:
            button_x = local_cursor_position.x() + horizontal_offset

        # Print the adjusted offsets and calculated button positions
        # print(f"Adjusted Offset: {offset}")

        # Ensure the button doesn't go beyond the webview bounds horizontally
        if local_cursor_position.x() + offset + button.width() > self.webview.width():
            offset = self.webview.width() - local_cursor_position.x() - button.width()

        # Ensure the button doesn't go out on the left side
        if local_cursor_position.x() + offset < 0:
            offset = -local_cursor_position.x()

        y_offset = 80  # default y offset moved down by 10

        # Ensure the button doesn't go beyond the webview bounds vertically
        if local_cursor_position.y() + y_offset + button.height() > self.webview.height():
            y_offset = self.webview.height() - local_cursor_position.y() - button.height()

        # Adjust for other buttons if they exist, ensuring they don't scatter
        button_gap = 1  # spacing between buttons

        # Default positioning for the button
        button_x = min(self.webview.width() - button.width(), max(0, local_cursor_position.x() + offset))
        button_y = min(self.webview.height() - button.height(), max(0, local_cursor_position.y() + y_offset))
        # print(f"Local Cursor Position: {local_cursor_position}")
        if button_key == "addReferenceButton":
            # Position Add Reference below Add Description if it exists, otherwise below the main button
            if "addDescriptionButton" in self.buttons:
                button_x = self.buttons["addDescriptionButton"].x()
                button_y = self.buttons["addDescriptionButton"].y() + self.buttons[
                    "addDescriptionButton"].height() + button_gap
            else:
                button_y = button_y + button.height() + button_gap

        elif button_key == "addDescriptionButton" and "addWordButton" in self.buttons:
            # Calculate desired position for Add Description next to Add Word
            desired_desc_x = self.buttons["addWordButton"].x() + self.buttons["addWordButton"].width() + button_gap
            # Check if Add Description goes beyond webview bounds
            if desired_desc_x + button.width() > self.webview.width():
                # Set Add Description to the maximum possible position without going out of bounds
                button_x = self.webview.width() - button.width()
                # Adjust Add Word position to make space for Add Description
                self.buttons["addWordButton"].move(button_x - self.buttons["addWordButton"].width() - button_gap,
                                                   button_y)
            else:
                button_x = desired_desc_x
            button_y = self.buttons["addWordButton"].y()

        elif button_key == "addWordButton" and "addDescriptionButton" in self.buttons:
            # Ensure Add Word is to the left of Add Description
            desc_button_new_x = button_x + button.width() + button_gap
            self.buttons["addDescriptionButton"].move(desc_button_new_x, button_y)

        # After positioning the buttons, check if they're still within the webview bounds
        button_x = min(self.webview.width() - button.width(), button_x)
        if "addDescriptionButton" in self.buttons and button_key == "addWordButton":
            desc_button_x = min(self.webview.width() - self.buttons["addDescriptionButton"].width(), desc_button_new_x)
            self.buttons["addDescriptionButton"].move(desc_button_x, button_y)

        # print(f"Final Button X: {button_x}")
        # print(f"Final Button Y: {button_y}")

        button.move(button_x, button_y)
        button.clicked.connect(callback)
        button.show()

        self.buttons[button_key] = button

    def add_selected_text(self):
        """
        DESCRIPTION: This function adds the selected text to the line edit.
        return:    None
        """
        selected_text = self.webview.selectedText()
        if selected_text:
            self.line_edit.setText(self.line_edit.text() + " " + selected_text)

    def get_current_page_number(self):
        """
        DESCRIPTION: This function gets the current page number.
        return:    None
        """
        worldId = 0  # default value
        self.webview.page().runJavaScript(
            "parseInt(PDFViewerApplication.page)", worldId, self.handle_page_number
        )

    def handle_page_number(self, result):
        try:
            # Ensure result is an integer.
            result = int(result)

            # If pre_page_edit is empty, set pre_page_value to 0
            if not self.pre_page_edit.text().strip():
                pre_page_value = 0
            else:
                try:
                    pre_page_value = int(self.pre_page_edit.text())
                except ValueError:
                    pre_page_value = 0
                    print("Please enter a valid integer for the number of pre-pages.")

            self.page_number = result - pre_page_value if result > pre_page_value else 0  # result

        except ValueError:
            self.page_number = result

    ## Here is for the search funtion
    def show_context_menu(self, position, table):
        self.contextMenuPos = position
        self.currentContextMenuTable = table

        # Create a RoundMenu instance and assign it to contextMenu.
        contextMenu = RoundMenu(table)

        # Add an action to the context menu, triggering self.add_selected_row_to_table when activated.
        addAction = Action(FIF.ADD, "Add selected row to table")  # Assuming FIF.ADD is an icon you have defined.
        addAction.triggered.connect(self.add_selected_row_to_table)
        contextMenu.addAction(addAction)

        # Display the context menu at the provided position.
        contextMenu.exec(table.viewport().mapToGlobal(position))

    def add_selected_row_to_table(self):
        # print("here 1")

        # Directly get the global position from the QCursor
        globalPos = QCursor.pos()

        # Convert the global position to the current table's local coordinates
        pos = self.currentContextMenuTable.viewport().mapFromGlobal(globalPos)

        # Find the row and column at the specified position
        row = self.currentContextMenuTable.rowAt(pos.y())
        col = self.currentContextMenuTable.columnAt(pos.x())

        print(f"Row: {row}, Column: {col}")

        # Ensure we got a valid row and column
        if row != -1 and col != -1:
            # print("here 2")
            # Get the data from the current table
            word_item = self.currentContextMenuTable.item(row, 0)
            page_item = self.currentContextMenuTable.item(row, 1)
            sentence_widget = self.currentContextMenuTable.item(row,
                                                                2)  # self.currentContextMenuTable.cellWidget(row, 2)

            if word_item and page_item and sentence_widget:
                # print("here 3")
                word = word_item.text()
                page = page_item.text()
                sentence = sentence_widget.text()
                self.add_text_from_search(word, page, sentence)
            else:
                print(f"Missing items at Row: {row}")
                if not sentence_widget:
                    print(f"No sentence widget for Row: {row}, Column: 2")
        self.currentContextMenuTable = None

    def apply_dialog_stylesheet(self, widget):
        if self.current_theme == Theme.DARK:
            widget.setStyleSheet("""
                QDialog, QMessageBox, QLabel {
                    background-color: #2b2b2b;
                    color: #cccccc;  /* Ensure text is light in dark mode */
                }
                QPushButton {
                    background-color: #333333;
                    color: #ffffff;
                }
                QLineEdit {
                    background-color: #333333;
                    color: #ffffff;
                }
                QSplitter::handle {
                    background-color: #2b2b2b;
                }
                QSplitter::handle:horizontal {
                    height: 1px;
                }
                QSplitter::handle:vertical {
                    width: 1px;
                }
                /* Additional styles for dark theme */
            """)
        else:
            widget.setStyleSheet("""
                QDialog, QMessageBox, QLabel {
                    background-color: #ffffff;
                    color: #000000;  /* Ensure text is dark in light mode */
                }
                QPushButton {
                    background-color: #e6e6e6;
                    color: #000000;
                }
                QLineEdit {
                    background-color: #ffffff;
                    color: #000000;
                }
                QSplitter::handle {
                    background-color: #ffffff;
                }
                QSplitter::handle:horizontal {
                    height: 1px;
                }
                QSplitter::handle:vertical {
                    width: 1px;
                }
                /* Additional styles for light theme */
            """)

    def display_search_results(self, results):

        # Check the current theme
        current_theme = self.current_theme
        # print(f"Current Theme: {current_theme}")

        # If the current theme is dark, set to light
        # if current_theme == Theme.DARK:
        # setTheme(Theme.LIGHT)
        # Update the combo box style in PDFIndexCreator
        # self.indexCreator.reapply_combo_box_style(Theme.LIGHT)

        # Delete existing buttons
        for key in list(self.buttons.keys()):
            self.buttons[key].deleteLater()
            self.buttons.pop(key)

        dialog = QDialog(self)
        # Check the current theme and apply the appropriate stylesheet
        self.apply_dialog_stylesheet(dialog)
        self.open_dialogs.append(dialog)  # Add the dialog to the list of open dialogs

        word = results[0][0] if results else ''
        dialog.setWindowTitle(f'Search Results for "{word}"')

        layout = QHBoxLayout()  # Change to QHBoxLayout

        # Create a new table for this search result
        table = TableWidget(self)

        # Apply the custom delegate to the required columns
        delegate = CustomTableItemDelegate(table, {TextEdit: [2], LineEdit: [0, 1]})
        table.setItemDelegate(delegate)

        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(['Word', 'Page', 'Sentence'])
        table.setRowCount(len(results))
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.setSortingEnabled(True)
        table.verticalHeader().hide()

        # Connect to the context menu using a lambda function to ensure proper signal-slot binding
        table.customContextMenuRequested.connect(lambda pos, t=table: self.show_context_menu(pos, t))

        for i, (word, page, sentence) in enumerate(results):
            normalized_word = ' '.join(word.split())  # Normalize the word to put it on one line
            table.setItem(i, 0, QTableWidgetItem(normalized_word))
            table.setItem(i, 1, QTableWidgetItem(str(page)))
            table.setItem(i, 2, QTableWidgetItem(sentence))

        table.horizontalHeader().setStretchLastSection(True)
        table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        table.resizeColumnToContents(0)
        table.resizeColumnToContents(1)

        # Connect the item clicked signal to a slot
        table.itemClicked.connect(self.show_page_from_search)

        # Create a PDF display widget on the right side
        # pdf_display = QWebEngineView()  # local variable
        pdf_display = QWebEngineView(dialog)
        pdf_display.load(QUrl.fromLocalFile(self.current_pdf_path))

        # Create a splitter
        splitter = QSplitter(Qt.Horizontal)

        # Add the table and the PDF display to the splitter
        splitter.addWidget(table)
        splitter.addWidget(pdf_display)

        # Set equal initial sizes for both widgets in the splitter
        splitter.setSizes([600, 600])

        # Set the splitter as the main widget for the dialog
        layout = QVBoxLayout(dialog)  # Create a main layout for the dialog
        layout.addWidget(splitter)  # Add the splitter to this layout
        dialog.setLayout(layout)
        dialog.resize(1800, 700)  # Resize the dialog to fit both the table and the PDF display

        # Store this table and its associated PDF display in a list (create them if they don't exist)
        if not hasattr(self, "search_tables"):
            self.search_tables = []
        if not hasattr(self, "pdf_displays"):
            self.pdf_displays = []
        self.search_tables.append(table)
        self.pdf_displays.append(pdf_display)
        # setTheme(Theme.LIGHT)
        # Hide the splitter in the main window when showing the dialog
        self.toggle_splitter_position()
        self.pdfVisibilityAction.setChecked(False)  # Uncheck the button

        # Reset the theme when the dialog is closed
        def dialog_finished():
            if dialog in self.open_dialogs:
                self.open_dialogs.remove(dialog)

        # Connect the finished signal of the dialog
        dialog.finished.connect(dialog_finished)

        dialog.show()

        def reset_theme():
            current_theme_now = self.current_theme

            # print(f"Current Theme: {current_theme_now} {current_theme}")
            if current_theme_now == current_theme:
                setTheme(current_theme)
                # Update the combo box style in PDFIndexCreator
                self.indexCreator.reapply_combo_box_style(current_theme)
            else:
                self.indexCreator.reapply_combo_box_style(current_theme_now)
            # Update the combo box style in PDFIndexCreator
            # self.indexCreator.reapply_combo_box_style(self.current_theme)
            current_text = self.pdfVisibilityAction.text()
            if current_text != 'Hide PDF Viewer':
                self.toggle_splitter_position()
                # self.pdfVisibilityAction.setText('Hide PDF Viewer')
                self.pdfVisibilityAction.setChecked(True)

        dialog.finished.connect(lambda: reset_theme())

        # dialog.accepted.connect(reset_theme)
        # dialog.rejected.connect(reset_theme)

    def show_page_from_search(self, item):
        table = item.tableWidget()  # Get the table that contains the clicked item
        index = self.search_tables.index(table)  # Find the index of this table in search_tables
        pdf_display = self.pdf_displays[index]  # Get the associated PDF display

        row = item.row()

        # Get the page number and word from the table
        page_number = int(table.item(row, 1).text())
        word = ' '.join(table.item(row, 0).text().split())

        # Adjust the page number using the pre_page_value
        if not self.pre_page_edit.text().strip():
            pre_page_value = 0
        else:
            try:
                pre_page_value = int(self.pre_page_edit.text())
            except ValueError:
                pre_page_value = 0
                print("Please enter a valid integer for the number of pre-pages.")
        adjusted_page_number = page_number + pre_page_value

        # Ensure JavaScript is enabled
        settings = pdf_display.page().settings()
        settings.setAttribute(QWebEngineSettings.JavascriptEnabled, True)

        # Only load the PDF page if it's different from the current one
        current_url = pdf_display.url().toString()
        target_url = f"{PDFJS}?file={self.current_pdf_path}#page={adjusted_page_number}"
        if current_url != target_url:
            pdf_display.load(QUrl(target_url))
            QTimer.singleShot(1000, lambda: self.select_word_in_pdf(word, pdf_display))
        else:
            self.select_word_in_pdf(word, pdf_display)

    def select_word_in_pdf(self, word, pdf_display):
        # print(f"Attempting to select and highlight all occurrences of the word: {word}...")

        # JavaScript code using PDFViewerApplication.eventBus.dispatch('find', {...})
        js_code = f"""
        // This method uses the viewer's search functionality to highlight given text
        if (typeof PDFViewerApplication !== 'undefined') {{
            PDFViewerApplication.eventBus.dispatch('find', {{
                query: '{word}',
                caseSensitive: false,
                highlightAll: true,
                findPrevious: false
            }});
        }}

        """

        # Execute the JavaScript code and get the diagnostic message
        pdf_display.page().runJavaScript(js_code)

    def get_search_words(self, initial_word=""):
        dialog = QDialog(self)
        self.apply_dialog_stylesheet(dialog)
        dialog.setWindowTitle("Enter Words to Search")

        layout = QVBoxLayout()
        splitter = QSplitter(Qt.Vertical)

        # Text area for user input
        textarea = TextEdit()
        textarea.setText(initial_word)  # Set the text of the text area to the initial word

        # Notes label
        notes_label = CaptionLabel('To search for multiple words, enter each word on a new line.\n'
                                   'For example:\n'
                                   'FirstWord\n'
                                   'SecondWord\n'
                                   'Tip: Make sure you have a correct pre-page value set', self)

        splitter.addWidget(textarea)
        splitter.addWidget(notes_label)

        button = PushButton("Search")

        layout.addWidget(splitter)
        layout.addWidget(button)
        dialog.setLayout(layout)

        words = []

        def collect_words():
            text = textarea.toPlainText().strip()
            if not text:
                InfoBar.error(
                    title='No Search Term',
                    content="Please enter a search term.",
                    orient=Qt.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=2000,
                    parent=self
                )
                return  # Exit the function early if no words were entered

            nonlocal words
            words_dict = {}
            for word in text.splitlines():
                stripped_word = word.strip()
                lowercase_word = stripped_word.lower()
                if lowercase_word not in words_dict:
                    words_dict[lowercase_word] = stripped_word

            words = list(words_dict.values())
            dialog.accept()

        button.clicked.connect(collect_words)
        dialog.exec()
        return words

    def search_in_pdf(self):
        # Check if a PDF is loaded
        if not hasattr(self, 'current_pdf_path') or not self.current_pdf_path:
            InfoBar.error(
                title='No PDF Loaded',
                content="Please load a PDF file before searching.",
                orient=Qt.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self
            )
            return

        # Clear the selection and remove buttons immediately.
        # self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")
        for key in list(self.buttons.keys()):  # make a copy of the keys
            self.buttons[key].deleteLater()
            self.buttons.pop(key)
        # self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")

        # Use QTimer to delay the dialog appearance, allowing GUI to update.
        QTimer.singleShot(20, self.show_search_dialog)

    def show_search_dialog(self):
        word = re.sub(r'\s+', ' ', self.line_edit.text().replace('\n', ' ')).strip()
        self.webview.page().runJavaScript("window.getSelection().removeAllRanges();")
        words = self.get_search_words(word)  # Pass the word from the line edit
        if not words:
            return

        all_results = []
        for word in words:
            self.proceed_with_search(word, accumulate_results=all_results)

        # After accumulating all the results, we display them
        self.display_search_results(all_results)

    def proceed_with_search(self, word, accumulate_results=None):

        # Checking the pre_page_edit value before searching
        if not self.pre_page_edit.text().strip():
            pre_page_value = 0
            filter_zero_page = False  # No need to filter out zero-page results
        else:
            try:
                pre_page_value = int(self.pre_page_edit.text())
                filter_zero_page = True  # We'll filter out zero-page results later
            except ValueError:
                pre_page_value = 0
                filter_zero_page = False
                print("Please enter a valid integer for the number of pre-pages.")

        with open(self.current_pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            results = []

            # Normalize the search word: replace newlines and multiple spaces
            normalized_word = re.sub(r'\s+', ' ', word).strip().lower()  # Convert to lowercase

            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                content = page.extract_text()

                # Normalize content: replace newlines and collapse multiple spaces into one
                normalized_content = re.sub(r'\s+', ' ', content).lower()  # Convert to lowercase

                if normalized_word in normalized_content:
                    for sentence in re.split('\.\s', normalized_content):
                        if normalized_word in sentence:
                            # Convert newline characters to HTML line break tags for rich text display
                            # sentence = sentence.replace('\n', '<br>')
                            # Adjust page number using the pre_page_value
                            if page_num < pre_page_value:
                                displayed_page_number = 0
                            else:
                                displayed_page_number = page_num + 1 - pre_page_value
                            results.append((word, displayed_page_number, sentence))

            # Filter out results with page number 0 if filter_zero_page is True
            if filter_zero_page:
                results = [result for result in results if result[1] != 0]

            # If accumulate_results is provided, append these results to it.
            if accumulate_results is not None:
                accumulate_results.extend(results)
            else:
                self.display_search_results(results)

    def add_text_from_search(self, word, page, sentence):
        """
        DESCRIPTION: This function adds the selected text from search results to the table.
        return:        None
        """
        book_number = self.book_edit.text().strip()

        # Check if word and page already exist in the table
        for row in range(self.indexCreator.table.rowCount()):
            existing_word_item = self.indexCreator.table.item(row, 0)
            existing_page_item = self.indexCreator.table.item(row, 1)
            existing_sentence_item = self.indexCreator.table.item(row, 2)

            # If word and page match, append sentence to the existing sentence
            if existing_word_item and existing_page_item:
                if existing_word_item.text() == word and existing_page_item.text() == page:
                    new_sentence = existing_sentence_item.text() + "\n" + sentence
                    self.indexCreator.table.setItem(row, 2, QTableWidgetItem(new_sentence))
                    return

        # If word and page do not exist, add a new row to the table
        row_position = self.indexCreator.table.rowCount()
        self.indexCreator.table.insertRow(row_position)
        self.indexCreator.table.setItem(row_position, 0, QTableWidgetItem(word))
        self.indexCreator.table.setItem(row_position, 1, QTableWidgetItem(page))
        self.indexCreator.table.setItem(row_position, 2, QTableWidgetItem(sentence))

        if book_number:
            self.indexCreator.table.setItem(row_position, 3, QTableWidgetItem(book_number))
        else:
            self.indexCreator.table.setItem(row_position, 3, QTableWidgetItem(" "))
        self.indexCreator.table.setItem(row_position, 4, QTableWidgetItem(" "))
        self.indexCreator.table.setItem(row_position, 5, QTableWidgetItem("None"))

        # Set the current row
        self.current_row = row_position

        if self.button is not None:
            self.button.deleteLater()
            self.button = None

        self.selection_state = SelectionState.WORD_SELECTED

        # Clean up and set some attributes
        for key in list(self.buttons.keys()):  # make a copy of the keys
            self.buttons[key].deleteLater()
            self.buttons.pop(key)

        self.indexCreator.table.clearSelection()

        # Set the selection state to DESCRIPTION_SELECTION
        self.selection_state = SelectionState.DESCRIPTION_SELECTION

        # Populate the "Icon" column for the newly added row
        self.indexCreator.populate_icon_column(row_position)

        # Save the state after a word is added
        self.indexCreator.save_state()


if __name__ == "__main__":
    """
    DESCRIPTION: This function is the main function.
    """

    setTheme(Theme.LIGHT)
    app = QApplication(sys.argv)
    app.setAttribute(Qt.AA_DontCreateNativeWidgetSiblings)
    # app.setWindowIcon(QIcon('icon.png'))
    SplashScreenWindow = SplashScreenWindow()  # replace MainWindow with SplashScreenWindow
    SplashScreenWindow.show()
    sys.exit(app.exec())

